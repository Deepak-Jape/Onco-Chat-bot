"""Generic {columns, data} -> .xlsx bytes, with optional per-cell comments and
an optional "Source Evidence" sheet.

Exists because CSV can't carry comments/notes: openpyxl writes an actual
.xlsx workbook and attaches an Excel "Comment" (the classic hover note, not
Excel's newer separate "threaded comment" feature -- this is the one every
Excel/LibreOffice/Google Sheets version renders on hover) to whichever cells
have one. Cells without a comment are written exactly as before -- there is
no fallback representation, they're just plain cells.

A comment alone is a poor fit for a full source excerpt: comments are
plain-text-only (see investigation notes in traceability.py -- Excel/openpyxl
cannot do partial-text highlighting or rich formatting inside a comment, and
tables can't be represented in one at all), and a long excerpt in a small
hover popup is cramped. So when `evidence` is supplied, the data cell keeps
its short comment for an at-a-glance summary AND gets a hyperlink to a
dedicated "Source Evidence" sheet holding the full excerpt/reasoning/
confidence/method, with its own hyperlink back out to the original source
URL and one back to the data cell -- click-through instead of a cramped
popup, using only standard, non-fragile Excel features (comments, cell
hyperlinks, a second worksheet).

Kept independent of any one table's shape: callers pass the same
{columns: [{key,label}], data: [...]} shape already used for the chart
blocks (see PanelTable.jsx / CohortTable.jsx), plus comments/evidence dicts.
"""

from io import BytesIO

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_COMMENT_AUTHOR = "OncoSuite"
# openpyxl / Excel silently truncates far past this, but keep comments well
# under the ~32k hard limit so they stay skimmable in the hover popup.
_MAX_COMMENT_CHARS = 1000

_DATA_SHEET = "Data"
_EVIDENCE_SHEET = "Source Evidence"
_HYPERLINK_FONT = Font(color="0563C1", underline="single")

# Header styling for both sheets -- solid fill + white bold text, standard
# Excel header convention, applied uniformly so the two sheets read as one
# workbook rather than the plain-white-with-bold-text header used before.
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F4E79")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

# Status badges: light fill + matching darker font, same green/amber logic
# tokens.js's statusColor() already uses in the web app -- but NOT its red
# for Terminated/Withdrawn/Suspended (explicit request: no red anywhere in
# this workbook). A calm slate/gray reads as "stopped" without the alarm
# connotation red carries, and stays visually distinct from the amber
# "pending" tier so the two aren't confused for the same meaning.
#
# Order matters: checked top-to-bottom, first match wins. "Not Yet
# Recruiting" contains the substring "recruit", so the amber "pending" tier
# MUST be checked before the green "in progress" catch-all, or every
# not-yet-started trial would incorrectly show as green -- tokens.js's own
# statusColor() has the same explicit ordering guard for this exact reason.
_STATUS_BADGES = (
    (("suspend", "withdraw", "terminat"), ("EDEFF2", "525A66")),  # slate: stopped -- no red
    (("complet",), ("E3EEFC", "1955A6")),             # blue: finished, successful arc
    (("not yet", "enroll"), ("FDF3DA", "8A6D1D")),    # amber: pending start
    (("recruit", "active"), ("E6F4EA", "1E7A34")),    # green: in progress (checked LAST)
)

# Confidence tiers on the Evidence sheet -- same visual language as status
# (green = strong, amber = moderate, slate = weak) instead of red for "low".
_CONFIDENCE_BADGES = (
    (90, ("E6F4EA", "1E7A34")),
    (70, ("FDF3DA", "8A6D1D")),
    (0, ("EDEFF2", "525A66")),
)


def _badge_for_status(status):
    v = (status or "").lower()
    for needles, colors in _STATUS_BADGES:
        if any(n in v for n in needles):
            return colors
    return None


def _badge_for_confidence(pct):
    if pct is None:
        return None
    for threshold, colors in _CONFIDENCE_BADGES:
        if pct >= threshold:
            return colors
    return None


def _apply_badge(cell, colors, bold=False):
    if not colors:
        return
    fill_color, font_color = colors
    cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
    cell.font = Font(color=font_color, bold=bold)
# Alternating shade per DATA ROW's block of evidence -- makes it obvious at a
# glance where one row's evidence ends and the next begins, without reading
# every "Row" label.
_GROUP_FILLS = (
    PatternFill(fill_type="solid", fgColor="FFFFFF"),
    PatternFill(fill_type="solid", fgColor="F2F6FC"),
)
# A visible rule at the start of each new group, on top of the shading --
# the two together are unambiguous even at a quick scroll-past glance.
_GROUP_TOP_BORDER = Border(top=Side(style="medium", color="B7C6DE"))

# "Row" is prepended here (column 1) so every evidence row is legible on its
# own without cross-referencing coordinates -- which data row this came from,
# in plain terms, not just a cell reference.
_EVIDENCE_COLUMNS = [
    ("row_label", "Row", 26),
    ("field_label", "Field", 16),
    ("extracted_value", "Extracted Value", 24),
    ("full_evidence", "Full Evidence (source excerpt)", 60),
    ("reasoning", "Reasoning", 40),
    ("confidence", "Confidence", 12),
    ("extraction_method", "Method", 16),
    ("source", "Source", 20),
]
_EVIDENCE_ROW_HEIGHT = 60
# Columns that hold prose (source excerpt, reasoning) always wrap, even
# single-line paragraphs with no literal "\n" -- otherwise Excel clips the
# text at the next non-empty cell's boundary instead of showing it (this was
# the actual cause of the truncated-looking text: wrap only ever triggered
# on an embedded newline, never on a long single-line paragraph).
_ALWAYS_WRAP_EVIDENCE_COLUMNS = {"full_evidence", "reasoning"}


def _set_value(ws, row, col, value, wrap=None):
    if wrap is None:
        wrap = isinstance(value, str) and "\n" in value
    cell = ws.cell(row=row, column=col, value=value if value != "" else None)
    if wrap:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    return cell


def build_xlsx(columns, rows, comments=None, evidence=None):
    """columns: [{"key": ..., "label": ...}, ...]
    rows: [{col_key: value, ...}, ...]
    comments: {(row_index, col_key): "comment text"} -- row_index is 0-based,
    matching rows' own indexing. Missing/None entries mean no comment.
    evidence: {(row_index, col_key): [record, ...]} -- record is the dict
    shape from traceability.to_evidence_records() (extracted_value,
    full_evidence, reasoning, confidence_pct, extraction_method, source_type,
    source_name, source_link, field_label). When a key has records, that data
    cell gets a hyperlink to its row(s) on the Source Evidence sheet, in
    addition to any comment already on that cell -- Excel treats a cell's
    comment and hyperlink as independent properties, so both work together.

    Returns the workbook as raw bytes (.getvalue() of an in-memory buffer),
    ready to base64-encode or stream as a download.
    """
    comments = comments or {}
    evidence = evidence or {}

    wb = Workbook()
    ws = wb.active
    ws.title = _DATA_SHEET

    for col_i, col in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_i, value=col.get("label", col["key"]))
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    for row_i, row in enumerate(rows):
        for col_i, col in enumerate(columns, start=1):
            # A "187 (Planned)\n112 (Enrolled)"-style stacked cell (see
            # dashboard.py's _n()) should read as one wrapped cell, not a
            # literal backslash-n -- matches how the UI already renders it.
            cell = _set_value(ws, row_i + 2, col_i, row.get(col["key"]))

            # Status badge -- green/blue/amber/slate by meaning, matching the
            # same tiers the React CohortTable already uses (tokens.js's
            # statusColor()) EXCEPT its red for Terminated/Withdrawn/
            # Suspended, deliberately swapped for a calmer slate here.
            if col["key"] == "status":
                _apply_badge(cell, _badge_for_status(row.get(col["key"])), bold=True)

            text = comments.get((row_i, col["key"]))
            if text:
                comment = Comment(text[:_MAX_COMMENT_CHARS], _COMMENT_AUTHOR)
                comment.width = 320
                comment.height = 160
                ws.cell(row=row_i + 2, column=col_i).comment = comment

    for col_i, col in enumerate(columns, start=1):
        width = col.get("width")
        ws.column_dimensions[ws.cell(row=1, column=col_i).column_letter].width = (
            max(10, min(40, width / 7)) if width else 18
        )

    # Header stays visible while scrolling, and every column gets a filter
    # dropdown -- both standard, both make a long table far easier to work
    # with without changing how the data itself is structured.
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    if evidence:
        _add_evidence_sheet(wb, ws, columns, rows, evidence)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _row_label(row, columns, row_i):
    """Human-readable identifier for an evidence block -- whatever the
    table's own first column shows (OncoSuite ID, NCT id, ...), not a raw
    row index. Falls back to a plain row number if the first column has no
    value for this row."""
    id_col = columns[0]["key"] if columns else None
    identifier = row.get(id_col) if id_col else None
    return f"Row {row_i + 2}" + (f" — {identifier}" if identifier else "")


def _add_evidence_sheet(wb, data_ws, columns, rows, evidence):
    ev_ws = wb.create_sheet(_EVIDENCE_SHEET)

    for col_i, (_key, label, _width) in enumerate(_EVIDENCE_COLUMNS, start=1):
        cell = ev_ws.cell(row=1, column=col_i, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    back_col = len(_EVIDENCE_COLUMNS) + 1
    link_col = back_col + 1
    for c, label in ((back_col, "Back to Data"), (link_col, "Source Link")):
        cell = ev_ws.cell(row=1, column=c, value=label)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    for col_i, (_key, _label, width) in enumerate(_EVIDENCE_COLUMNS, start=1):
        ev_ws.column_dimensions[ev_ws.cell(row=1, column=col_i).column_letter].width = width
    ev_ws.column_dimensions[ev_ws.cell(row=1, column=back_col).column_letter].width = 18
    ev_ws.column_dimensions[ev_ws.cell(row=1, column=link_col).column_letter].width = 16

    # Column-key -> column-index lookup, so the hyperlink written back onto
    # the data cell lands on the exact cell the evidence came from.
    col_index = {col["key"]: i for i, col in enumerate(columns, start=1)}

    # Iterate in DATA-SHEET reading order (row-major), not dict order, so the
    # evidence sheet reads top-to-bottom the same way the data sheet does.
    next_row = 2
    group_index = 0
    max_row_i = max((k[0] for k in evidence), default=-1)
    for row_i in range(max_row_i + 1):
        row_has_evidence = any((row_i, col["key"]) in evidence for col in columns)
        if not row_has_evidence:
            continue
        row_label = _row_label(rows[row_i], columns, row_i)
        group_fill = _GROUP_FILLS[group_index % 2]
        group_index += 1
        group_start_row = next_row

        for col in columns:
            key = (row_i, col["key"])
            records = evidence.get(key)
            if not records:
                continue

            first_evidence_row = next_row
            data_coord = data_ws.cell(row=row_i + 2, column=col_index[col["key"]]).coordinate

            for record in records:
                # Per-record, not hoisted above this loop -- a single cell's
                # evidence list can combine more than one underlying field
                # (e.g. "indication" = Histology records + Line of Therapy
                # records; "n" = N Target + N Actual), so each record keeps
                # its OWN label rather than all of them showing the first
                # record's label.
                field_label = record.get("field_label") or col.get("label")
                source_bits = [b for b in (record.get("source_type"), record.get("source_name")) if b]
                conf = record.get("confidence_pct")
                ev_ws.row_dimensions[next_row].height = _EVIDENCE_ROW_HEIGHT
                ev_ws.cell(row=next_row, column=1, value=row_label)
                ev_ws.cell(row=next_row, column=2, value=field_label)
                ev_ws.cell(row=next_row, column=3, value=record.get("extracted_value"))
                _set_value(ev_ws, next_row, 4, record.get("full_evidence"), wrap=True)
                _set_value(ev_ws, next_row, 5, record.get("reasoning"), wrap=True)
                ev_ws.cell(row=next_row, column=6, value=f"{conf}%" if conf is not None else None)
                ev_ws.cell(row=next_row, column=7, value=record.get("extraction_method"))
                ev_ws.cell(row=next_row, column=8, value=" / ".join(source_bits) or None)

                back_cell = ev_ws.cell(
                    row=next_row, column=back_col, value=f"↩ Row {row_i + 2} · {field_label}"
                )
                back_cell.hyperlink = f"#'{_DATA_SHEET}'!{data_coord}"
                back_cell.font = _HYPERLINK_FONT

                link = record.get("source_link")
                if link:
                    link_cell = ev_ws.cell(row=next_row, column=link_col, value="View Source")
                    link_cell.hyperlink = link
                    link_cell.font = _HYPERLINK_FONT

                # Shade every cell in this data-row's block (not just the
                # ones with an explicit value) so the banding reads as one
                # continuous block rather than speckled cells. Also top-align
                # EVERY column, not just the two that opt into wrap_text
                # above -- the row is 60px tall to fit wrapped prose, and a
                # short value (e.g. "Phase", "100%") left at Excel's default
                # bottom alignment sits at the bottom of that tall row while
                # the wrapped columns sit at the top, so the row visually
                # splits into two misaligned halves instead of reading as
                # one row.
                for c in range(1, link_col + 1):
                    cell = ev_ws.cell(row=next_row, column=c)
                    cell.fill = group_fill
                    cell.alignment = Alignment(vertical="top", wrap_text=True)

                # Confidence badge -- applied AFTER the group shading above,
                # so it isn't overwritten by it. Same green/amber/slate tiers
                # as the status badges (no red for low confidence either).
                conf_cell = ev_ws.cell(row=next_row, column=6)
                _apply_badge(conf_cell, _badge_for_confidence(conf), bold=True)
                conf_cell.alignment = Alignment(vertical="top", wrap_text=True)

                next_row += 1

            data_cell = data_ws.cell(row=row_i + 2, column=col_index[col["key"]])
            data_cell.hyperlink = f"#'{_EVIDENCE_SHEET}'!A{first_evidence_row}"
            # Keep the cell's own font as-is (bold headers etc. are on row 1
            # only, so this never fires there) except for tinting it like a
            # standard Excel hyperlink -- a visual cue that this cell has
            # full evidence a click away, independent of its comment.
            existing = data_cell.font
            data_cell.font = Font(
                color=_HYPERLINK_FONT.color, underline="single",
                bold=existing.bold if existing else None,
            )

        # A visible rule where this data row's block starts -- on top of the
        # alternating shade, so the boundary is unambiguous even scrolling
        # past quickly, not just on close inspection of the fill color.
        for c in range(1, link_col + 1):
            cell = ev_ws.cell(row=group_start_row, column=c)
            cell.border = Border(top=_GROUP_TOP_BORDER.top)

    ev_ws.freeze_panes = "A2"
    ev_ws.auto_filter.ref = ev_ws.dimensions
